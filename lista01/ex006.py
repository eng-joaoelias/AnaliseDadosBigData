import pandas as pd
import openpyxl

def escreveArquivo(char_init, char_final, num_init, num_final):
    df = pd.DataFrame()

    char_ascii_inicio = ord(char_init)
    char_ascii_fim = ord(char_final)

    for linha in range(num_init, num_final + 1):
        for char_ascii in range(ord(char_init), ord(char_final) + 1):
            coluna = chr(char_ascii)
            stringColunaLinha = coluna + str(linha)
            df.loc[linha - num_init + 2, coluna] = stringColunaLinha

    # Salvar o DataFrame como um arquivo Excel
    nome_arquivo = "escrever_linhas_colunas.xlsx"
    df.to_excel(nome_arquivo, index=False)

    # Exibir o conteúdo do arquivo Excel no terminal
    wb = openpyxl.load_workbook(nome_arquivo)
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        print(f"Conteúdo do arquivo '{nome_arquivo}', Planilha: '{sheet_name}':")
        for row in sheet.iter_rows(values_only=True):
            print("\t".join(str(cell) for cell in row))

def main():
    col_inicio = input("Escreva a coluna de início>> ").upper()
    col_fim = input("Escreva a coluna de fim>> ").upper()
    linha_inicio = int(input("Entre com a linha de inicio>> "))
    linha_fim = int(input("Entre com a linha de fim>> "))
    escreveArquivo(col_inicio, col_fim, linha_inicio, linha_fim)

if __name__ == "__main__":
    main()
